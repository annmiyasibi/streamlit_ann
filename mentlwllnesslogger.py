import tkinter as tk
from tkinter import ttk, messagebox
import pandas as pd
import re
import os
import platform
import subprocess
from datetime import datetime
import threading
from openpyxl import load_workbook, Workbook

class MentalWellnessLogger:
    def __init__(self, root):
        self.root = root
        self.root.title("Mental Wellness Entry Logger")
        self.root.geometry("800x600")
        self.root.configure(bg="#fdfdfd")

        dashboard = tk.Frame(root, bg="#eae6f8", height=60)
        dashboard.pack(fill='x')
        tk.Label(
            dashboard,
            text="Mental Wellness Entry Logger Dashboard",
            font=("Segoe UI", 16, "bold"),
            fg="#4b0082",
            bg="#eae6f8"
        ).pack(pady=10)

        content_frame = tk.Frame(root, bg="#fdfdfd", padx=20, pady=20)
        content_frame.pack(fill='both', expand=True)

        left_frame = tk.LabelFrame(
            content_frame, text="Log Your Wellness",
            bg="#ffffff", fg="#4b0082",
            font=("Segoe UI", 12, "bold"),
            padx=15, pady=15, bd=2, relief='groove'
        )
        left_frame.grid(row=0, column=0, sticky='n', padx=10, pady=10)

        tk.Label(left_frame, text="Student Name:", font=("Segoe UI", 10), bg="#ffffff").grid(row=0, column=0, sticky='w', pady=5)
        self.name_entry = ttk.Entry(left_frame, width=30)
        self.name_entry.grid(row=0, column=1, pady=5)

        tk.Label(left_frame, text="Mental Wellness Activity:", font=("Segoe UI", 10), bg="#ffffff").grid(row=1, column=0, sticky='w', pady=5)
        self.activity_entry = ttk.Entry(left_frame, width=30)
        self.activity_entry.grid(row=1, column=1, pady=5)

        tk.Label(left_frame, text="Me-Time Activity:", font=("Segoe UI", 10), bg="#ffffff").grid(row=2, column=0, sticky='w', pady=5)
        self.me_time_entry = ttk.Entry(left_frame, width=30)
        self.me_time_entry.grid(row=2, column=1, pady=5)

        tk.Label(left_frame, text="Screen-Free Time (minutes):", font=("Segoe UI", 10), bg="#ffffff").grid(row=3, column=0, sticky='w', pady=5)
        self.screen_time_entry = ttk.Entry(left_frame, width=30)
        self.screen_time_entry.grid(row=3, column=1, pady=5)
        self.screen_time_entry.bind("<KeyRelease>", lambda e: self.update_status())

        btn_frame = tk.Frame(left_frame, bg="#ffffff")
        btn_frame.grid(row=4, column=0, columnspan=2, pady=15)

        self.make_button(btn_frame, "Add Entry", "#bcaedc", self.add_entry).pack(side='left', padx=5)
        self.make_button(btn_frame, "Delete Selected", "#bcaedc", self.delete_entry).pack(side='left', padx=5)
        self.make_button(btn_frame, "Clear All", "#bcaedc", self.clear_all).pack(side='left', padx=5)
        self.make_button(btn_frame, "Save to Excel", "#bcaedc", self.save_to_excel).pack(side='left', padx=5)
        self.make_button(btn_frame, "Clear Form", "#bcaedc", self.clear_inputs).pack(side='left', padx=5)

        right_frame = tk.LabelFrame(
            content_frame, text="Logged Entries",
            bg="#ffffff", fg="#4b0082",
            font=("Segoe UI", 12, "bold"),
            padx=15, pady=15, bd=2, relief='groove'
        )
        right_frame.grid(row=0, column=1, sticky='nsew', padx=10, pady=10)

        self.entry_listbox = tk.Listbox(
            right_frame, width=60, height=15,
            bg="#f5f4fb", fg="#4b0082", font=("Consolas", 10)
        )
        self.entry_listbox.pack(pady=5)

        status_frame = tk.Frame(root, bg="#eae6f8", height=40)
        status_frame.pack(fill='x', side='bottom')
        tk.Label(status_frame, text="Wellness Status:", bg="#eae6f8", fg="#4b0082", font=("Segoe UI", 11)).pack(side='left', padx=10)
        self.status_var = tk.StringVar(value="Healthy")
        self.status_label = tk.Label(status_frame, textvariable=self.status_var, bg="#eae6f8", fg="green", font=("Segoe UI", 11, "bold"))
        self.status_label.pack(side='left')

        self.entries = []
        self.reminder_popup()  # Start reminder loop

    def make_button(self, parent, text, color, command):
        return tk.Button(parent, text=text, bg=color, fg="white", font=("Segoe UI", 9, "bold"), command=command)

    def is_valid_text(self, text):
        return re.fullmatch(r"[A-Za-z\s]+", text) is not None

    def update_status(self, *args):
        screen_time = self.screen_time_entry.get().strip()
        if not screen_time.isdigit() or int(screen_time) <= 0:
            self.status_var.set("Needs More Me-Time")
            self.status_label.config(fg='red')
            return
        screen_time_int = int(screen_time)
        if screen_time_int >= 120:
            self.status_var.set("Healthy")
            self.status_label.config(fg='green')
        else:
            self.status_var.set("Needs More Me-Time")
            self.status_label.config(fg='red')

    def add_entry(self):
        name = self.name_entry.get().strip()
        activity = self.activity_entry.get().strip()
        me_time = self.me_time_entry.get().strip()
        screen_time = self.screen_time_entry.get().strip()

        if not name or not activity or not me_time or not screen_time:
            messagebox.showerror("Error", "All fields are required.")
            return
        if not self.is_valid_text(name):
            messagebox.showerror("Error", "Name must contain only letters and spaces.")
            return
        if not self.is_valid_text(activity):
            messagebox.showerror("Error", "Mental Wellness Activity must contain only letters and spaces.")
            return
        if not self.is_valid_text(me_time):
            messagebox.showerror("Error", "Me-Time Activity must contain only letters and spaces.")
            return
        if not screen_time.isdigit() or int(screen_time) <= 0:
            messagebox.showerror("Error", "Screen-free time must be a positive number.")
            return

        screen_time_int = int(screen_time)
        self.update_status()

        entry_str = f"{name} | {activity} | {me_time} | {screen_time} mins"
        self.entry_listbox.insert(tk.END, entry_str)

        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        status = self.status_var.get()
        new_entry = {
            "Timestamp": timestamp,
            "Name": name,
            "Wellness Activity": activity,
            "Me-Time Activity": me_time,
            "Off-screen Time (min)": screen_time_int,
            "Frequency": "1",
            "Status": status
        }
        self.entries.append(new_entry)

        file_path = "class_wellness_data.xlsx"
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["Timestamp", "Name", "Wellness Activity", "Me-Time Activity", "Off-screen Time (min)", "Frequency", "Status"])

        ws.append(list(new_entry.values()))
        wb.save(file_path)

        messagebox.showinfo("Success", "Record added successfully.")
        self.clear_inputs()

    def delete_entry(self):
        selected = self.entry_listbox.curselection()
        if not selected:
            messagebox.showerror("Error", "No entry selected to delete.")
            return
        index = selected[0]
        self.entry_listbox.delete(index)
        del self.entries[index]
        messagebox.showinfo("Deleted", "Record deleted successfully.")

    def clear_all(self):
        self.entry_listbox.delete(0, tk.END)
        self.entries.clear()
        messagebox.showinfo("Cleared", "All records cleared.")

    def clear_inputs(self):
        self.name_entry.delete(0, tk.END)
        self.activity_entry.delete(0, tk.END)
        self.me_time_entry.delete(0, tk.END)
        self.screen_time_entry.delete(0, tk.END)
        self.status_var.set("Healthy")
        self.status_label.config(fg='green')

    def save_to_excel(self):
        if not self.entries:
            messagebox.showerror("Error", "No records to save.")
            return
        df = pd.DataFrame(self.entries)
        file_path = "mental_wellness_log.xlsx"
        try:
            df.to_excel(file_path, index=False)
            messagebox.showinfo("Saved", f"Records saved to Excel file:\n{file_path}")
            self.open_excel_file(file_path)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save to Excel.\n{e}")

    def open_excel_file(self, path):
        try:
            if platform.system() == 'Windows':
                os.startfile(path)
            elif platform.system() == 'Darwin':
                subprocess.run(['open', path])
            else: 
                subprocess.run(['xdg-open', path])
        except Exception as e:
            messagebox.showwarning("Open File", f"File saved, but couldn't open automatically.\n{e}")

    def reminder_popup(self):
        messagebox.showinfo("Reminder", "Take a mental wellness break!\nLog your me-time or screen-free activity.")
        threading.Timer(28800, self.reminder_popup).start()  # every 8 hours

if __name__ == "__main__":
    root = tk.Tk()
    app = MentalWellnessLogger(root)
    root.mainloop()
